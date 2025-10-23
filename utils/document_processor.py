import os
import json
import PyPDF2
from docx import Document
import openpyxl
import pdf2image
import pytesseract
from PIL import Image
import rispy
import re
from typing import Dict, List, Any
import subprocess
import tempfile

from config.settings import settings


class DocumentProcessor:
    def __init__(self):
        self.supported_formats = settings.SUPPORTED_EXTENSIONS

    def process_document(self, file_path: str) -> Dict[str, Any]:
        """处理单个文档，返回标准化的JSON结构"""
        file_ext = os.path.splitext(file_path)[1].lower()

        if file_ext not in self.supported_formats:
            raise ValueError(f"不支持的文档格式: {file_ext}")

        processors = {
            '.txt': self._process_txt,
            '.md': self._process_md,
            '.docx': self._process_docx,
            '.pptx': self._process_pptx,
            '.xlsx': self._process_xlsx,
            '.pdf': self._process_pdf,
            '.caj': self._process_caj,
            '.enw': self._process_enw,
            '.ris': self._process_ris,
            '.tex': self._process_tex,
            '.jpg': self._process_image,
            '.png': self._process_image
        }

        return processors[file_ext](file_path)

    def _extract_text_from_scanned_pdf(self, pdf_path: str) -> str:
        """从扫描版PDF提取文本"""
        try:
            images = pdf2image.convert_from_path(pdf_path)
            text = ""
            for image in images:
                text += pytesseract.image_to_string(image, lang='chi_sim+eng')
            return text
        except Exception as e:
            print(f"OCR处理失败: {e}")
            return ""

    def _process_txt(self, file_path: str) -> Dict[str, Any]:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return self._build_json_structure(file_path, content)

    def _process_md(self, file_path: str) -> Dict[str, Any]:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return self._build_json_structure(file_path, content)

    def _process_docx(self, file_path: str) -> Dict[str, Any]:
        doc = Document(file_path)
        content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return self._build_json_structure(file_path, content)

    def _process_pdf(self, file_path: str) -> Dict[str, Any]:
        text = ""
        try:
            # 首先尝试直接提取文本
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"

            # 如果文本太少，可能是扫描件，使用OCR
            if len(text.strip()) < 100:
                ocr_text = self._extract_text_from_scanned_pdf(file_path)
                if ocr_text:
                    text = ocr_text

        except Exception as e:
            print(f"PDF处理错误: {e}")
            text = self._extract_text_from_scanned_pdf(file_path)

        return self._build_json_structure(file_path, text)

    def _process_caj(self, file_path: str) -> Dict[str, Any]:
        """处理CAJ文件：先转换为PDF再处理"""
        try:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
                # 使用caj2pdf转换（需要安装caj2pdf）
                subprocess.run(['caj2pdf', 'convert', file_path, '-o', temp_pdf.name],
                               check=True, capture_output=True)
                return self._process_pdf(temp_pdf.name)
        except Exception as e:
            print(f"CAJ转换失败: {e}")
            return self._build_json_structure(file_path, "")

    def _process_enw(self, file_path: str) -> Dict[str, Any]:
        """处理ENW文件（EndNote）"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 简单的ENW解析
        structured_info = {}
        lines = content.split('\n')
        current_field = None

        for line in lines:
            if line.startswith('%'):
                parts = line.split(' ', 1)
                if len(parts) > 1:
                    field = parts[0][1:]
                    value = parts[1].strip()
                    structured_info[field] = value

        return {
            "title": structured_info.get('T', ''),
            "content": content,
            "structured_info": structured_info,
            "format_source": "ENW",
            "file_path": file_path
        }

    def _process_ris(self, file_path: str) -> Dict[str, Any]:
        """处理RIS文件"""
        with open(file_path, 'r', encoding='utf-8') as f:
            entries = rispy.load(f)

        if entries:
            entry = entries[0]
            return {
                "title": entry.get('title', ''),
                "content": str(entry),
                "structured_info": entry,
                "format_source": "RIS",
                "file_path": file_path
            }
        return self._build_json_structure(file_path, "")

    def _process_pptx(self, file_path: str) -> Dict[str, Any]:
        # 简化处理，实际需要python-pptx库
        return self._build_json_structure(file_path, "PPTX内容提取")

    def _process_xlsx(self, file_path: str) -> Dict[str, Any]:
        workbook = openpyxl.load_workbook(file_path)
        content = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content += f"【表格】工作表: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                content += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
            content += "\n"
        return self._build_json_structure(file_path, content)

    def _process_tex(self, file_path: str) -> Dict[str, Any]:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return self._build_json_structure(file_path, content)

    def _process_image(self, file_path: str) -> Dict[str, Any]:
        """处理图片文件"""
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            return self._build_json_structure(file_path, text)
        except Exception as e:
            print(f"图片处理失败: {e}")
            return self._build_json_structure(file_path, "")

    def _build_json_structure(self, file_path: str, content: str) -> Dict[str, Any]:
        """构建标准化的JSON结构"""
        # 清理冗余信息
        cleaned_content = self._clean_content(content)

        # 提取段落
        paragraphs = self._split_paragraphs(cleaned_content)

        # 识别格式标签
        formatted_content = self._add_format_tags(paragraphs)

        return {
            "title": os.path.basename(file_path),
            "content": formatted_content,
            "structured_info": {
                "file_type": os.path.splitext(file_path)[1],
                "file_size": os.path.getsize(file_path),
                "processing_time": "2024-01-01"  # 实际应该用当前时间
            },
            "format_source": os.path.splitext(file_path)[1].upper().replace('.', ''),
            "file_path": file_path,
            "paragraphs": paragraphs
        }

    def _clean_content(self, content: str) -> str:
        """清理冗余信息"""
        # 移除过多的空白字符
        content = re.sub(r'\s+', ' ', content)
        # 移除页眉页脚等常见噪音（根据实际情况调整）
        content = re.sub(r'第\s*\d+\s*页', '', content)
        return content.strip()

    def _split_paragraphs(self, content: str) -> List[str]:
        """将内容分割为段落"""
        paragraphs = re.split(r'\n\s*\n', content)
        return [p.strip() for p in paragraphs if p.strip()]

    def _add_format_tags(self, paragraphs: List[str]) -> str:
        """添加格式标签"""
        tagged_content = []
        for para in paragraphs:
            # 检测表格（简单的启发式规则）
            if re.search(r'\|\s*.+\s*\|', para) or re.search(r'\+[-]+\+', para):
                tagged_content.append(f"【表格】{para}")
            # 检测公式（简单的启发式规则）
            elif re.search(r'\$[^$]+\$|\\\(.*\\\)|\\\[.*\\\]', para):
                tagged_content.append(f"【公式】{para}")
            else:
                tagged_content.append(para)
        return "\n\n".join(tagged_content)